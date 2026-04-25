"""
Rentab v0.3 — формирование Excel-сметы как документа.

Функция generate_estimate_xlsx собирает .xlsx из результата расчёта
(результат страницы «Проект» — st.session_state["results"]) и возвращает
его байтами для st.download_button. Файл всегда содержит два листа:

  1) «Смета»            — клиентский документ. Содержимое и колонки
     зависят от модели ценообразования: для биллинга — детализация
     по часам и ставкам, для фикс-прайса — укрупнённые позиции без
     раскрытия ставок и часов.
  2) «Внутренний расчёт» — технический breakdown для фирмы: издержки,
     налог, NNE, маржа; для биллинга — Blended Rate / Leverage,
     для фикс-прайса — анализ этапов со светофором.

Стилизация унифицирована — единые цвета и форматы заданы константами
в начале модуля; никаких «магических» чисел в логике сборки.
"""

from __future__ import annotations

import io
from datetime import date
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


# ---------------------------------------------------------------------------
# Стилевые константы.
# Все цвета указаны без знака # — таков формат openpyxl.styles.PatternFill.
# Один источник истины: меняем здесь — меняется во всех листах.
# ---------------------------------------------------------------------------
HEADER_FILL_COLOR: str = "1F497D"      # тёмно-синий — шапка документа
TABLE_HEADER_FILL_COLOR: str = "D9E1F2"  # светло-голубой — шапка таблицы

HEADER_FONT_COLOR: str = "FFFFFF"
DEFAULT_FONT_NAME: str = "Arial"

HEADER_FONT_SIZE: int = 12
DEFAULT_FONT_SIZE: int = 10

# Числовой формат: «1 234 567» (тысячный разделитель — пробел, без копеек).
NUMBER_FORMAT_INT: str = '# ##0'

# Запас ширины столбца сверх длины самой длинной ячейки (символов).
COLUMN_WIDTH_PADDING: int = 2

# Минимально и максимально допустимые ширины столбцов (символов).
COLUMN_WIDTH_MIN: int = 10
COLUMN_WIDTH_MAX: int = 60


# ---------------------------------------------------------------------------
# Шрифты и заливки — собираем один раз и переиспользуем.
# ---------------------------------------------------------------------------
def _header_fill() -> PatternFill:
    """Возвращает заливку для шапки документа (тёмно-синяя)."""
    return PatternFill(
        start_color=HEADER_FILL_COLOR,
        end_color=HEADER_FILL_COLOR,
        fill_type="solid",
    )


def _table_header_fill() -> PatternFill:
    """Возвращает заливку для шапки таблицы (светло-голубая)."""
    return PatternFill(
        start_color=TABLE_HEADER_FILL_COLOR,
        end_color=TABLE_HEADER_FILL_COLOR,
        fill_type="solid",
    )


def _header_font() -> Font:
    """Жирный белый Arial 12 — для заголовка документа."""
    return Font(
        name=DEFAULT_FONT_NAME,
        size=HEADER_FONT_SIZE,
        bold=True,
        color=HEADER_FONT_COLOR,
    )


def _table_header_font() -> Font:
    """Жирный чёрный Arial 10 — для шапки таблицы."""
    return Font(name=DEFAULT_FONT_NAME, size=DEFAULT_FONT_SIZE, bold=True)


def _default_font() -> Font:
    """Обычный Arial 10 — для тела таблицы."""
    return Font(name=DEFAULT_FONT_NAME, size=DEFAULT_FONT_SIZE)


def _bold_font() -> Font:
    """Жирный Arial 10 — для строк итогов."""
    return Font(name=DEFAULT_FONT_NAME, size=DEFAULT_FONT_SIZE, bold=True)


def _total_top_border() -> Border:
    """Тонкая верхняя граница — отделяет строку итогов от данных."""
    return Border(top=Side(style="thin", color="000000"))


# ---------------------------------------------------------------------------
# Низкоуровневые помощники для записи строк
# ---------------------------------------------------------------------------
def _write_document_header(
    ws: Worksheet,
    row: int,
    title: str,
    n_cols: int,
) -> int:
    """Пишет шапку-заголовок документа на строке row.

    Объединяет ячейки A{row}:{lastcol}{row}, заливает HEADER_FILL_COLOR,
    выводит белым жирным шрифтом строку title по центру.

    Args:
        ws: Лист openpyxl.
        row: Номер строки (1-based) для заголовка.
        title: Текст заголовка.
        n_cols: Сколько колонок объединить (ширина шапки).

    Returns:
        Номер следующей свободной строки.
    """
    last_col_letter = get_column_letter(n_cols)
    ws.merge_cells(f"A{row}:{last_col_letter}{row}")
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = _header_font()
    cell.fill = _header_fill()
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 24
    return row + 1


def _write_table_header(
    ws: Worksheet,
    row: int,
    headers: list[str],
) -> int:
    """Пишет шапку таблицы (жирные подписи колонок на голубом фоне)."""
    for col_idx, header_text in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header_text)
        cell.font = _table_header_font()
        cell.fill = _table_header_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center")
    return row + 1


def _write_data_row(
    ws: Worksheet,
    row: int,
    values: list[Any],
    number_columns: set[int] | None = None,
) -> int:
    """Пишет строку данных. Колонки из number_columns форматируются как числа."""
    if number_columns is None:
        number_columns = set()
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.font = _default_font()
        if col_idx in number_columns:
            cell.number_format = NUMBER_FORMAT_INT
            cell.alignment = Alignment(horizontal="right")
    return row + 1


def _write_total_row(
    ws: Worksheet,
    row: int,
    values: list[Any],
    number_columns: set[int] | None = None,
) -> int:
    """Пишет строку итогов: жирно + тонкая верхняя граница."""
    if number_columns is None:
        number_columns = set()
    border = _total_top_border()
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.font = _bold_font()
        cell.border = border
        if col_idx in number_columns:
            cell.number_format = NUMBER_FORMAT_INT
            cell.alignment = Alignment(horizontal="right")
    return row + 1


def _autosize_columns(ws: Worksheet, n_cols: int) -> None:
    """Подгоняет ширины колонок 1..n_cols под длину самой длинной ячейки.

    Объединённые ячейки игнорируются: их «контент» формально лежит только
    в верхней-левой ячейке диапазона, что искажало бы ширину под колонку
    шапки. Ограничиваем ширину коридором [COLUMN_WIDTH_MIN, COLUMN_WIDTH_MAX].
    """
    merged_top_left: set[str] = {
        str(rng).split(":")[0] for rng in ws.merged_cells.ranges
    }
    for col_idx in range(1, n_cols + 1):
        col_letter = get_column_letter(col_idx)
        max_len = COLUMN_WIDTH_MIN
        for cell in ws[col_letter]:
            # Пропускаем ячейки шапки документа (объединены).
            if cell.coordinate in merged_top_left:
                continue
            if cell.value is None:
                continue
            text = (
                f"{cell.value:,.0f}".replace(",", " ")
                if isinstance(cell.value, (int, float))
                else str(cell.value)
            )
            if len(text) > max_len:
                max_len = len(text)
        width = min(max_len + COLUMN_WIDTH_PADDING, COLUMN_WIDTH_MAX)
        ws.column_dimensions[col_letter].width = max(width, COLUMN_WIDTH_MIN)


# ---------------------------------------------------------------------------
# Сборка листа «Смета» (клиентский документ)
# ---------------------------------------------------------------------------
def _build_client_sheet_billing(
    ws: Worksheet,
    results: dict,
    firm_name: str,
) -> None:
    """Заполняет лист «Смета» для биллинговой модели.

    Колонки таблицы: №, Этап, Исполнитель, Часы, Ставка, Стоимость.
    Дополнительно — блок перевыставляемых расходов (если есть)
    и итоговая строка «Итого к оплате».
    """
    project_name = (results.get("project_name") or "Без названия").strip() or "Без названия"
    client = results.get("client", "") or ""
    today_str = date.today().strftime("%d.%m.%Y")
    currency_symbol = results.get("currency", "₽")

    headers = ["№", "Этап", "Исполнитель", "Часы", f"Ставка, {currency_symbol}/ч", f"Стоимость, {currency_symbol}"]
    n_cols = len(headers)

    row = 1
    row = _write_document_header(ws, row, f"Смета по проекту: {project_name}", n_cols)

    # Метаданные документа: фирма / клиент / дата.
    if firm_name:
        ws.cell(row=row, column=1, value=f"Исполнитель: {firm_name}").font = _default_font()
        row += 1
    if client:
        ws.cell(row=row, column=1, value=f"Заказчик: {client}").font = _default_font()
        row += 1
    ws.cell(row=row, column=1, value=f"Дата: {today_str}").font = _default_font()
    row += 1
    if results.get("regime_label"):
        ws.cell(
            row=row,
            column=1,
            value=f"Налоговый режим: {results['regime_label']}",
        ).font = _default_font()
        row += 1
    row += 1  # пустая строка-разделитель

    # Таблица работ.
    row = _write_table_header(ws, row, headers)
    number_cols = {4, 5, 6}

    team_rows = list(results.get("team_with_hours", []))
    sum_cost = 0.0
    for idx, m in enumerate(team_rows, start=1):
        hours = float(m.get("hours", 0.0))
        billing = float(m.get("billing_rate", 0.0))
        cost = billing * hours
        sum_cost += cost
        row = _write_data_row(
            ws,
            row,
            [
                idx,
                str(m.get("stage_name", "—")),
                f"{m.get('name', '—')} ({m.get('role', '—')})",
                hours,
                billing,
                cost,
            ],
            number_columns=number_cols,
        )

    row = _write_total_row(
        ws,
        row,
        ["", "Итого за услуги", "", "", "", sum_cost],
        number_columns=number_cols,
    )

    # Блок перевыставляемых расходов — только если они есть.
    disbursements_billed = float(results.get("disbursements", 0.0))
    if disbursements_billed > 0:
        row += 1
        ws.cell(row=row, column=1, value="Перевыставляемые расходы (пошлины и иные платежи)").font = _bold_font()
        row += 1
        billable_expenses = [
            e for e in results.get("project_expenses", [])
            if e.get("billable") and e.get("category") == "disbursement_billable"
        ]
        for idx, e in enumerate(billable_expenses, start=1):
            amount = float(e.get("amount", 0.0))
            row = _write_data_row(
                ws,
                row,
                [idx, str(e.get("name", "")), "", "", "", amount],
                number_columns=number_cols,
            )
        row = _write_total_row(
            ws,
            row,
            ["", "Итого перевыставляемых расходов", "", "", "", disbursements_billed],
            number_columns=number_cols,
        )

    # Финальная строка «Итого к оплате».
    total_client = float(results.get("total_client", sum_cost + disbursements_billed))
    row += 1
    row = _write_total_row(
        ws,
        row,
        ["", "Итого к оплате", "", "", "", total_client],
        number_columns=number_cols,
    )

    # Подвал.
    row += 2
    ws.cell(
        row=row,
        column=1,
        value="Смета действительна 30 дней с даты составления.",
    ).font = _default_font()

    _autosize_columns(ws, n_cols)


def _build_client_sheet_fixed(
    ws: Worksheet,
    results: dict,
    firm_name: str,
) -> None:
    """Заполняет лист «Смета» для фикс-прайс модели.

    Часы и ставки клиенту не показываем — фикс-прайс продаётся как
    результат, а не как time & material. Колонки: №, Описание, Стоимость.
    """
    project_name = (results.get("project_name") or "Без названия").strip() or "Без названия"
    client = results.get("client", "") or ""
    today_str = date.today().strftime("%d.%m.%Y")
    currency_symbol = results.get("currency", "₽")

    headers = ["№", "Описание", f"Стоимость, {currency_symbol}"]
    n_cols = len(headers)

    row = 1
    row = _write_document_header(ws, row, f"Смета по проекту: {project_name}", n_cols)

    if firm_name:
        ws.cell(row=row, column=1, value=f"Исполнитель: {firm_name}").font = _default_font()
        row += 1
    if client:
        ws.cell(row=row, column=1, value=f"Заказчик: {client}").font = _default_font()
        row += 1
    ws.cell(row=row, column=1, value=f"Дата: {today_str}").font = _default_font()
    row += 1
    if results.get("regime_label"):
        ws.cell(
            row=row,
            column=1,
            value=f"Налоговый режим: {results['regime_label']}",
        ).font = _default_font()
        row += 1
    row += 1

    # Таблица работ — этапы как самостоятельные строки документа.
    row = _write_table_header(ws, row, headers)
    number_cols = {3}

    fixed_price = float(results.get("fixed_price", 0.0))
    stage_flags = list(results.get("stage_flags", []))
    if stage_flags:
        for idx, sf in enumerate(stage_flags, start=1):
            row = _write_data_row(
                ws,
                row,
                [
                    idx,
                    str(sf.get("stage_name", "—")),
                    float(sf.get("stage_revenue_share", 0.0)),
                ],
                number_columns=number_cols,
            )
    else:
        # Если этапов нет — одна строка «Услуги по проекту».
        row = _write_data_row(
            ws,
            row,
            [1, "Услуги по проекту (фиксированная стоимость)", fixed_price],
            number_columns=number_cols,
        )

    row = _write_total_row(
        ws,
        row,
        ["", "Фиксированная стоимость работ", fixed_price],
        number_columns=number_cols,
    )

    # Сноска про фикс-прайс.
    row += 1
    ws.cell(
        row=row,
        column=1,
        value="* Стоимость работ зафиксирована и не зависит от фактически затраченных часов.",
    ).font = _default_font()
    row += 1

    # Перевыставляемые расходы (если применимы и в фикс-модели).
    disbursements_billed = float(results.get("disbursements", 0.0))
    if disbursements_billed > 0:
        row += 1
        ws.cell(
            row=row, column=1,
            value="Перевыставляемые расходы (пошлины и иные платежи)",
        ).font = _bold_font()
        row += 1
        billable_expenses = [
            e for e in results.get("project_expenses", [])
            if e.get("billable") and e.get("category") == "disbursement_billable"
        ]
        for idx, e in enumerate(billable_expenses, start=1):
            amount = float(e.get("amount", 0.0))
            row = _write_data_row(
                ws,
                row,
                [idx, str(e.get("name", "")), amount],
                number_columns=number_cols,
            )
        row = _write_total_row(
            ws,
            row,
            ["", "Итого перевыставляемых расходов", disbursements_billed],
            number_columns=number_cols,
        )

    total_client = float(
        results.get("total_client", fixed_price + disbursements_billed)
    )
    row += 1
    row = _write_total_row(
        ws,
        row,
        ["", "Итого к оплате", total_client],
        number_columns=number_cols,
    )

    row += 2
    ws.cell(
        row=row,
        column=1,
        value="Смета действительна 30 дней с даты составления.",
    ).font = _default_font()

    _autosize_columns(ws, n_cols)


# ---------------------------------------------------------------------------
# Сборка листа «Внутренний расчёт»
# ---------------------------------------------------------------------------
def _build_internal_sheet_billing(ws: Worksheet, results: dict) -> None:
    """Внутренний breakdown для биллинговой модели.

    Содержит финансовые показатели проекта: выручка, издержки, налог, NNE,
    маржа, Blended Rate, Leverage. Используется фирмой, не клиентом.
    """
    currency_symbol = results.get("currency", "₽")
    n_cols = 2
    headers = ["Показатель", "Значение"]

    row = 1
    row = _write_document_header(ws, row, "Внутренний расчёт проекта", n_cols)
    row += 1

    row = _write_table_header(ws, row, headers)
    number_cols = {2}

    gross = float(results.get("gross_revenue", 0.0))
    nne = float(results.get("nne", 0.0))
    margin_pct = (nne / gross * 100.0) if gross else 0.0

    rows: list[tuple[str, Any, bool]] = [
        (f"Выручка за услуги, {currency_symbol}", gross, True),
        (f"Перевыставляемые расходы, {currency_symbol}",
         float(results.get("disbursements", 0.0)), True),
        (f"Итоговый счёт клиенту, {currency_symbol}",
         float(results.get("total_client", 0.0)), True),
        (f"Прямые трудозатраты, {currency_symbol}",
         float(results.get("direct_labor", 0.0)), True),
        (f"Накладные (распределённые), {currency_symbol}",
         float(results.get("overheads_alloc", 0.0)), True),
        (f"Расходы за счёт фирмы, {currency_symbol}",
         float(results.get("disbursements_own", 0.0)), True),
        (f"Налог, {currency_symbol}", float(results.get("tax", 0.0)), True),
        (f"Чистая прибыль (NNE), {currency_symbol}", nne, True),
        ("Маржа, %", round(margin_pct, 1), False),
        (f"Средневзвешенная ставка, {currency_symbol}/ч",
         float(results.get("blended_rate", 0.0)), True),
        ("Коэффициент рычага", round(float(results.get("leverage", 0.0)), 2), False),
        ("Налоговый режим", str(results.get("regime_label", "—")), False),
    ]
    for label, value, is_int in rows:
        if is_int:
            row = _write_data_row(ws, row, [label, value], number_columns=number_cols)
        else:
            # Не-целочисленные значения (проценты, текст) пишем без формата.
            ws.cell(row=row, column=1, value=label).font = _default_font()
            ws.cell(row=row, column=2, value=value).font = _default_font()
            row += 1

    _autosize_columns(ws, n_cols)


def _build_internal_sheet_fixed(ws: Worksheet, results: dict) -> None:
    """Внутренний breakdown для фикс-прайс модели.

    Помимо общих показателей включает анализ этапов со «светофором»:
    зелёный / жёлтый / красный по марже относительно целевой.
    """
    currency_symbol = results.get("currency", "₽")

    # --- блок 1: финансовые показатели ---
    n_cols_block1 = 2
    headers_block1 = ["Показатель", "Значение"]

    row = 1
    row = _write_document_header(
        ws,
        row,
        "Внутренний расчёт проекта (фикс-прайс)",
        n_cols_block1,
    )
    row += 1

    row = _write_table_header(ws, row, headers_block1)
    number_cols = {2}

    target_margin_pct = float(results.get("target_margin", 0.0)) * 100.0
    actual_margin_pct = float(results.get("actual_margin", 0.0)) * 100.0

    rows1: list[tuple[str, Any, bool]] = [
        (f"Фиксированная цена, {currency_symbol}",
         float(results.get("fixed_price", 0.0)), True),
        (f"Все издержки проекта, {currency_symbol}",
         float(results.get("total_costs", 0.0)), True),
        (f"Прямые трудозатраты, {currency_symbol}",
         float(results.get("direct_labor", 0.0)), True),
        (f"Накладные (распределённые), {currency_symbol}",
         float(results.get("overheads_alloc", 0.0)), True),
        (f"Расходы за счёт фирмы, {currency_symbol}",
         float(results.get("disbursements_own", 0.0)), True),
        (f"Налог, {currency_symbol}", float(results.get("tax", 0.0)), True),
        (f"Чистая прибыль (NNE), {currency_symbol}",
         float(results.get("nne", 0.0)), True),
        ("Целевая маржа, %", round(target_margin_pct, 1), False),
        ("Фактическая маржа, %", round(actual_margin_pct, 1), False),
        ("Налоговый режим", str(results.get("regime_label", "—")), False),
    ]
    for label, value, is_int in rows1:
        if is_int:
            row = _write_data_row(ws, row, [label, value], number_columns=number_cols)
        else:
            ws.cell(row=row, column=1, value=label).font = _default_font()
            ws.cell(row=row, column=2, value=value).font = _default_font()
            row += 1

    # --- блок 2: анализ этапов со светофором ---
    stage_flags = list(results.get("stage_flags", []))
    if stage_flags:
        row += 2
        ws.cell(row=row, column=1, value="Анализ этапов").font = _bold_font()
        row += 1

        flag_labels = {"green": "Норма", "yellow": "Низкая маржа", "red": "Убыточный"}
        flag_fills = {
            "green":  PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid"),
            "yellow": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
            "red":    PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
        }

        headers_block2 = [
            "Этап",
            f"Издержки этапа, {currency_symbol}",
            f"Доля выручки, {currency_symbol}",
            "Маржа, %",
            "Статус",
        ]
        n_cols_block2 = len(headers_block2)
        row = _write_table_header(ws, row, headers_block2)

        for sf in stage_flags:
            flag_value = str(sf.get("flag", "green"))
            margin_pct = float(sf.get("stage_margin", 0.0)) * 100.0
            values = [
                str(sf.get("stage_name", "—")),
                float(sf.get("stage_costs", 0.0)),
                float(sf.get("stage_revenue_share", 0.0)),
                round(margin_pct, 1),
                flag_labels.get(flag_value, flag_value),
            ]
            row_fill = flag_fills.get(flag_value)
            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.font = _default_font()
                if row_fill is not None:
                    cell.fill = row_fill
                if col_idx in (2, 3):
                    cell.number_format = NUMBER_FORMAT_INT
                    cell.alignment = Alignment(horizontal="right")
            row += 1

        _autosize_columns(ws, max(n_cols_block1, n_cols_block2))
    else:
        _autosize_columns(ws, n_cols_block1)


# ---------------------------------------------------------------------------
# Публичный API модуля
# ---------------------------------------------------------------------------
def generate_estimate_xlsx(results: dict, firm_name: str = "") -> bytes:
    """Собирает .xlsx-смету из словаря results и возвращает байты файла.

    Файл всегда содержит два листа: «Смета» (для клиента) и «Внутренний
    расчёт» (для фирмы). Вид каждого листа зависит от results["pricing_model"]:
      - "billing" — детализация по часам/ставкам;
      - "fixed"   — укрупнённые позиции без раскрытия часов.

    Args:
        results: Результат расчёта со страницы «Проект» (st.session_state["results"]).
            Обязан содержать как минимум pricing_model и финансовые ключи
            соответствующей модели.
        firm_name: Название фирмы-исполнителя для шапки сметы. Если пусто,
            строка «Исполнитель» в шапке не выводится.

    Returns:
        Байты .xlsx-файла, готовые к передаче в st.download_button.
    """
    pricing_model = str(results.get("pricing_model", "billing"))

    wb = Workbook()
    ws_client = wb.active
    ws_client.title = "Смета"
    ws_internal = wb.create_sheet("Внутренний расчёт")

    if pricing_model == "fixed":
        _build_client_sheet_fixed(ws_client, results, firm_name)
        _build_internal_sheet_fixed(ws_internal, results)
    else:
        _build_client_sheet_billing(ws_client, results, firm_name)
        _build_internal_sheet_billing(ws_internal, results)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
