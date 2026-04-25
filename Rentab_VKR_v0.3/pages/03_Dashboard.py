"""
Rentab v0.3 — страница 03: Дашборд.

Страница рассчитана на показ уже сформированной на «Проекте» сметы
(st.session_state["results"]). Структура зависит от выбранной модели
ценообразования:

— pricing_model == "billing" (поведение v0.2):
    1. Карточки верхнего уровня — выручка, перевыставляемые расходы,
       итоговый счёт клиенту, чистая прибыль (NNE) с маржой.
    2. Структура цены (plotly pie) + вспомогательные метрики справа.
    3. Таблица по этапам с итоговой строкой.
    4. Таблица расходов проекта (показывается только если есть).
    5. Индикатор рентабельности + выгрузка в Excel.

— pricing_model == "fixed" (v0.3):
    1. Карточки: все издержки / целевая маржа / фикс. цена / NNE+реальная маржа.
    2. Горизонтальный stacked bar — из чего складывается цена клиенту:
       Прямые трудозатраты | Накладные | Налог | Прибыль.
    3. Таблица «Анализ этапов» со светофором (green / yellow / red).
    4. Индикатор рентабельности.

Для юрисдикции «Оба» доступна конвертация валюты в сайдбаре: пользователь
вводит курс RUB/KZT и выбирает целевую валюту — все суммы пересчитываются.

Все строки интерфейса переведены на русский. Ключи результата (gross_revenue,
direct_labor, fixed_price и т.п.) остаются английскими — они внутренние,
часть API между страницами.
"""

from __future__ import annotations

import io
from datetime import date

import pandas as pd
import plotly.graph_objects as go
import streamlit as st

from modules.jurisdiction import CURRENCY_CODES, CURRENCY_SYMBOLS


# ---------------------------------------------------------------------------
# Подписи статусов этапов (для фикс-прайс анализа).
# Используются в колонке «Статус» таблицы анализа этапов.
# ---------------------------------------------------------------------------
STAGE_FLAG_LABELS: dict[str, str] = {
    "green":  "🟢 Норма",
    "yellow": "🟡 Низкая маржа",
    "red":    "🔴 Убыточный",
}

# Цвета фоновой подсветки строк по флагу. Пустая строка означает «без подсветки».
STAGE_FLAG_BG_COLORS: dict[str, str] = {
    "green":  "",
    "yellow": "background-color: #fef3c7",   # пастельно-жёлтый
    "red":    "background-color: #fee2e2",   # пастельно-красный
}


# ---------------------------------------------------------------------------
# Справочник русских подписей для категорий расходов Expense
# ---------------------------------------------------------------------------
EXPENSE_CATEGORY_LABELS: dict[str, str] = {
    "overhead": "Накладные фирмы",
    "disbursement_billable": "Перевыставляемый клиенту",
    "disbursement_own": "За счёт фирмы",
    "project_extra": "Прочие проектные расходы",
}


# ---------------------------------------------------------------------------
# Заголовок
# ---------------------------------------------------------------------------
st.title("📊 Дашборд")

# ---------------------------------------------------------------------------
# Проверка: смета сформирована?
# ---------------------------------------------------------------------------
# Итерация 3: основной ключ — "results". На старых сессиях (до переименования)
# данные могут лежать в "project_result" — fallback сохраняем для совместимости.
r: dict | None = st.session_state.get("results") or st.session_state.get("project_result")
if r is None:
    st.info("Сначала сформируйте смету на странице «Проект»")
    st.stop()


# ---------------------------------------------------------------------------
# Сайдбар: конвертация валюты для юрисдикции «Оба»
# ---------------------------------------------------------------------------
# Логика: результат посчитан в валюте calc_country (например RUB). Если
# пользователь выбрал «Оба», даём на дашборде отдельно переключить валюту
# отображения и пересчитать по курсу. Множитель conv_factor применяется
# ко всем денежным величинам; символ валюты — соответствующий.
source_code: str = r.get("currency_code", "RUB")
source_symbol: str = r.get("currency", CURRENCY_SYMBOLS.get(source_code[:2], "₽"))

display_code = source_code
display_symbol = source_symbol
conv_factor = 1.0

if r.get("jurisdiction_mode") == "Both":
    with st.sidebar:
        st.subheader("💱 Валюта отображения")
        stored_rate = float(
            st.session_state.get("exchange_rate_rub_per_kzt", 0.18)
        )
        rate_input = st.number_input(
            "Курс: 1 KZT = ? RUB",
            min_value=0.0001,
            value=stored_rate,
            step=0.01,
            format="%.4f",
            help="Например, 0.18 ₽/₸ означает «1 ₸ ≈ 0,18 ₽».",
            key="dash_exchange_rate",
        )
        target_code = st.radio(
            "Показать суммы в",
            options=["RUB", "KZT"],
            index=0 if source_code == "RUB" else 1,
            horizontal=True,
            key="dash_target_currency",
        )
        if st.button("🔄 Пересчитать", width='stretch'):
            st.session_state["display_currency_code"] = target_code
            st.session_state["exchange_rate_rub_per_kzt"] = rate_input

    display_code = st.session_state.get("display_currency_code", source_code)
    rate_used = float(
        st.session_state.get("exchange_rate_rub_per_kzt", stored_rate)
    )

    # Считаем коэффициент перевода из source_code в display_code:
    # 1 KZT = rate_used RUB  →  1 RUB = 1/rate_used KZT.
    if display_code == source_code:
        conv_factor = 1.0
    elif source_code == "RUB" and display_code == "KZT":
        conv_factor = 1.0 / rate_used if rate_used > 0 else 1.0
    elif source_code == "KZT" and display_code == "RUB":
        conv_factor = rate_used
    # CURRENCY_SYMBOLS индексируется КОДАМИ СТРАН (RF/KZ), а display_code —
    # ISO-код валюты (RUB/KZT). Чтобы не путать сущности, держим локальный
    # словарь ISO → символ.
    display_symbol = {"RUB": "₽", "KZT": "₸"}.get(display_code, source_symbol)


def cx(value: float) -> float:
    """Применяет коэффициент конвертации валют к денежной величине."""
    return float(value) * conv_factor


sym = display_symbol  # короткий алиас для вёрстки

# ---------------------------------------------------------------------------
# Идентификация проекта
# ---------------------------------------------------------------------------
st.subheader(r.get("project_name") or "Без названия")
meta_parts: list[str] = []
if r.get("client"):
    meta_parts.append(f"**Клиент:** {r['client']}")
if r.get("regime_label"):
    meta_parts.append(f"**Режим:** {r['regime_label']}")
if meta_parts:
    st.caption("  ·  ".join(meta_parts))

# ===========================================================================
# Ветка для фикс-прайс модели — рендерит свой набор блоков и завершает
# выполнение скрипта через st.stop(). Биллинговая ветка ниже остаётся
# нетронутой и идентична v0.2.
# ===========================================================================
if r.get("pricing_model") == "fixed":
    # ---------- 1. Карточки верхнего уровня (4 шт.) -------------------- #
    direct_labor_x = cx(r["direct_labor"])
    overheads_alloc_x = cx(r["overheads_alloc"])
    disbursements_own_x = cx(r["disbursements_own"])
    tax_x = cx(r["tax"])
    fixed_price_x = cx(r["fixed_price"])
    total_costs_x = cx(r["total_costs"])
    nne_fp_x = cx(r["nne"])
    target_margin_pct = float(r.get("target_margin", 0.0)) * 100.0
    actual_margin_pct = float(r.get("actual_margin", 0.0)) * 100.0

    fp_c1, fp_c2, fp_c3, fp_c4 = st.columns(4)
    fp_c1.metric(f"Все издержки проекта, {sym}", f"{total_costs_x:,.0f}")
    fp_c2.metric("Целевая маржа", f"{target_margin_pct:.0f}%")
    fp_c3.metric(f"Фиксированная цена, {sym}", f"{fixed_price_x:,.0f}")
    fp_c4.metric(
        f"Чистая прибыль (NNE), {sym}",
        f"{nne_fp_x:,.0f}",
        delta=f"{actual_margin_pct:.1f}% реальная маржа",
    )

    st.markdown("---")

    # ---------- 2. Горизонтальный stacked bar — структура цены --------- #
    # Слои: Прямые трудозатраты | Накладные | Налог | Прибыль.
    # disbursements_own (если есть) включаем в «Накладные», чтобы получить
    # ровно четыре слоя как требует методика ВКР; сумма всех слоёв равна
    # fixed_price точно (Total Costs + Tax + NNE = Fixed Price).
    st.markdown("#### Структура фиксированной цены")

    indirect_x = overheads_alloc_x + disbursements_own_x

    bar_layers: list[tuple[str, float, str]] = [
        ("Прямые трудозатраты", direct_labor_x, "#3498db"),
        ("Накладные",           indirect_x,     "#9b59b6"),
        ("Налог",               tax_x,          "#e74c3c"),
        ("Прибыль",             nne_fp_x,       "#2ecc71"),
    ]

    fig_fp = go.Figure()
    for layer_name, layer_value, layer_color in bar_layers:
        fig_fp.add_trace(
            go.Bar(
                x=[layer_value],
                y=["Фикс. цена"],
                name=layer_name,
                orientation="h",
                marker=dict(color=layer_color),
                hovertemplate=(
                    f"<b>{layer_name}</b>: %{{x:,.0f}} {sym}<extra></extra>"
                ),
            )
        )
    fig_fp.update_layout(
        barmode="stack",
        height=200,
        margin=dict(t=10, b=10, l=10, r=10),
        legend=dict(orientation="h", yanchor="top", y=-0.3),
        xaxis=dict(title=f"Сумма, {sym}", tickformat=",.0f"),
        yaxis=dict(showticklabels=False),
    )
    st.plotly_chart(fig_fp, width='stretch')

    st.caption(f"Налоговый режим: {r.get('regime_label', '—')}")

    st.markdown("---")

    # ---------- 3. Таблица «Анализ этапов» с светофором ---------------- #
    st.markdown("#### Анализ этапов")

    # Часы по этапам соберём из team_with_hours, группируя по stage_name.
    # FixedPriceCalculator.get_stage_flags не возвращает часы, чтобы остаться
    # в рамках спецификации, поэтому считаем здесь — единый источник часов
    # лежит в team_with_hours (его страница «Проект» собирает всегда).
    hours_by_stage: dict[str, float] = {}
    for member_row in r.get("team_with_hours", []):
        stage_label = str(member_row.get("stage_name", ""))
        hours_by_stage[stage_label] = (
            hours_by_stage.get(stage_label, 0.0)
            + float(member_row.get("hours", 0.0))
        )

    flag_rows: list[dict] = []
    for stage_flag in r.get("stage_flags", []):
        stage_label = str(stage_flag.get("stage_name", "—"))
        flag_value = str(stage_flag.get("flag", "green"))
        flag_rows.append(
            {
                "Этап": stage_label,
                "Часы": f"{hours_by_stage.get(stage_label, 0.0):,.1f}",
                f"Издержки этапа, {sym}": f"{cx(stage_flag.get('stage_costs', 0.0)):,.0f}",
                f"Доля выручки, {sym}": f"{cx(stage_flag.get('stage_revenue_share', 0.0)):,.0f}",
                "Маржа, %": f"{float(stage_flag.get('stage_margin', 0.0)) * 100:.1f}",
                "Статус": STAGE_FLAG_LABELS.get(flag_value, flag_value),
                # Служебная колонка для подсветки строк — скрываем при выводе.
                "_flag": flag_value,
            }
        )

    if flag_rows:
        flag_df = pd.DataFrame(flag_rows)

        def _color_by_flag(row: pd.Series) -> list[str]:
            """Подсвечивает строку фоном по значению служебной колонки _flag."""
            bg_style = STAGE_FLAG_BG_COLORS.get(row.get("_flag", ""), "")
            return [bg_style for _ in row]

        styled_flags = flag_df.style.apply(_color_by_flag, axis=1)
        # Скрываем служебный _flag — он нужен только для подсветки.
        styled_flags = styled_flags.hide(subset=["_flag"], axis="columns")
        st.dataframe(styled_flags, width='stretch', hide_index=True)
    else:
        st.info("Нет этапов для анализа.")

    st.markdown("---")

    # ---------- 4. Индикатор рентабельности ---------------------------- #
    nne_raw_fp = float(r["nne"])
    if nne_raw_fp > 0:
        st.success("✅ Проект рентабелен")
    elif nne_raw_fp == 0:
        st.warning("⚠️ Проект в точке безубыточности")
    else:
        st.error(
            "❌ Проект убыточен — пересмотрите целевую маржу или часы по этапам."
        )

    # Завершаем рендер дашборда: дальше идёт billing-ветка v0.2,
    # которая обращается к gross_revenue и т.п. — для фикс-прайса этих
    # ключей нет, поэтому останавливаемся.
    st.stop()


# ---------------------------------------------------------------------------
# Блок 1 — Карточки верхнего уровня (биллинговая модель)
# ---------------------------------------------------------------------------
gross = cx(r["gross_revenue"])
disb_b = cx(r["disbursements"])
total_client = cx(r.get("total_client", r["gross_revenue"] + r["disbursements"]))
nne_val = cx(r["nne"])

# Перевыставляемые расходы показываем только если они ненулевые:
# для чистого «услуги без пошлин» колонку можно спрятать.
show_disb_col = disb_b > 0

if show_disb_col:
    c1, c2, c3, c4 = st.columns(4)
    c1.metric(f"Выручка (без перевыст.), {sym}", f"{gross:,.0f}")
    c2.metric(f"Перевыставляемые расходы, {sym}", f"{disb_b:,.0f}")
    c3.metric(f"Итоговый счёт клиенту, {sym}", f"{total_client:,.0f}")
    c4.metric(
        f"Чистая прибыль (NNE), {sym}",
        f"{nne_val:,.0f}",
        delta=(
            # Маржа PSF считается от ВЫРУЧКИ ЗА УСЛУГИ (gross), а не от
            # итогового счёта клиенту: пошлины — транзитные, прибыли не несут.
            f"{(nne_val / gross * 100):.1f}% маржи"
            if gross
            else None
        ),
    )
else:
    c1, c2, c3 = st.columns(3)
    c1.metric(f"Выручка, {sym}", f"{gross:,.0f}")
    c2.metric(f"Итоговый счёт клиенту, {sym}", f"{total_client:,.0f}")
    c3.metric(
        f"Чистая прибыль (NNE), {sym}",
        f"{nne_val:,.0f}",
        delta=(
            # Маржа PSF считается от ВЫРУЧКИ ЗА УСЛУГИ (gross), а не от
            # итогового счёта клиенту: пошлины — транзитные, прибыли не несут.
            f"{(nne_val / gross * 100):.1f}% маржи"
            if gross
            else None
        ),
    )

st.markdown("---")


# ---------------------------------------------------------------------------
# Блок 2 — Структура цены + вспомогательные метрики
# ---------------------------------------------------------------------------
col_chart, col_metrics = st.columns([1.4, 1])

with col_chart:
    st.markdown("#### Структура цены")

    # Секции в рублёвом эквиваленте и с русскими подписями.
    # Ноль и отрицательные значения в диаграмму не включаем.
    chart_sections: list[tuple[str, float, str]] = [
        ("Прямые трудозатраты",    cx(r["direct_labor"]),        "#3498db"),
        ("Накладные (распределённые)", cx(r["overheads_alloc"]), "#9b59b6"),
        ("Налог",                  cx(r["tax"]),                 "#e74c3c"),
        ("Расходы за счёт фирмы",  cx(r["disbursements_own"]),   "#95a5a6"),
        ("Чистая прибыль (NNE)",   nne_val,                      "#2ecc71"),
    ]
    labels = [name for name, val, _ in chart_sections if val > 0]
    values = [val for _, val, _ in chart_sections if val > 0]
    colors = [clr for _, val, clr in chart_sections if val > 0]

    if values:
        fig = go.Figure(
            data=[
                go.Pie(
                    labels=labels,
                    values=values,
                    marker=dict(colors=colors),
                    hole=0.35,
                    textinfo="label+percent",
                    hovertemplate="%{label}: %{value:,.0f} " + sym + "<extra></extra>",
                )
            ]
        )
        fig.update_layout(
            margin=dict(t=10, b=10, l=10, r=10),
            height=400,
            showlegend=True,
            legend=dict(orientation="h", yanchor="top", y=-0.05),
        )
        st.plotly_chart(fig, width='stretch')
    else:
        st.info("Все статьи нулевые — структуру цены построить нельзя.")

    st.caption(f"Налоговый режим: {r.get('regime_label', '—')}")

with col_metrics:
    st.markdown("#### Показатели")

    overhead_rate_val = float(
        r.get("overhead_rate", st.session_state.get("overhead_rate", 0.0))
    )

    # Эффективная ставка = налог / выручка (с учётом перевыставляемых — это
    # «итоговый счёт клиенту»; именно на эту сумму клиент смотрит).
    effective_rate = 0.0
    if total_client > 0:
        effective_rate = cx(r["tax"]) / total_client * 100

    st.metric(
        f"Средневзвешенная ставка, {sym}/ч",
        f"{cx(r['blended_rate']):,.0f}",
        help="Σ(billing × hours) / Σ(hours). С учётом валюты отображения.",
    )
    st.metric(
        "Коэффициент рычага",
        f"{r['leverage']:.2f}",
        help="Часы младших / часы старших. Выше — выше маржинальность.",
    )
    st.metric(
        f"Ставка накладных, {sym}/ч",
        f"{cx(overhead_rate_val):,.1f}",
        help="Σ(месячных накладных) / billable_hours_per_month.",
    )
    st.metric(
        "Эффективная ставка налога",
        f"{effective_rate:.1f}%",
        help=f"Налог / итоговый счёт. Базовый режим: {r.get('regime_label', '—')}.",
    )

st.markdown("---")


# ---------------------------------------------------------------------------
# Блок 3 — Таблица по этапам
# ---------------------------------------------------------------------------
st.markdown("#### Этапы проекта")

team_rows: list[dict] = list(r.get("team_with_hours", []))

if not team_rows:
    st.info("В смете нет этапов для отображения.")
    stages_df = pd.DataFrame()
else:
    stage_rows: list[dict] = []
    sum_hours = 0.0
    sum_revenue = 0.0
    sum_cost = 0.0
    for row in team_rows:
        hours = float(row.get("hours", 0.0))
        billing = float(row.get("billing_rate", 0.0))
        cost = float(row.get("cost_rate", 0.0))
        revenue = cx(billing * hours)
        labor_cost = cx(cost * hours)
        margin_pct = ((revenue - labor_cost) / revenue * 100) if revenue > 0 else 0.0
        stage_rows.append(
            {
                "Этап": row.get("stage_name", "—"),
                "Исполнитель": f"{row.get('name', '—')} ({row.get('role', '—')})",
                "Часы": f"{hours:,.1f}",
                f"Ставка ({sym}/ч)": f"{cx(billing):,.0f}",
                f"Выручка этапа, {sym}": f"{revenue:,.0f}",
                "Маржа, %": f"{margin_pct:.1f}",
            }
        )
        sum_hours += hours
        sum_revenue += revenue
        sum_cost += labor_cost

    total_margin_pct = ((sum_revenue - sum_cost) / sum_revenue * 100) if sum_revenue > 0 else 0.0
    stage_rows.append(
        {
            "Этап": "Итого",
            "Исполнитель": "",
            "Часы": f"{sum_hours:,.1f}",
            f"Ставка ({sym}/ч)": "",
            f"Выручка этапа, {sym}": f"{sum_revenue:,.0f}",
            "Маржа, %": f"{total_margin_pct:.1f}",
        }
    )
    stages_df = pd.DataFrame(stage_rows)

    # Жирная строка итогов через pandas Styler (st.dataframe поддерживает Styler).
    def _bold_last(row: pd.Series) -> list[str]:
        """Возвращает CSS-стили для строки: жирный шрифт для строки «Итого»."""
        is_total = row["Этап"] == "Итого"
        return ["font-weight: 700" if is_total else "" for _ in row]

    styled = stages_df.style.apply(_bold_last, axis=1)
    st.dataframe(styled, width='stretch', hide_index=True)


# ---------------------------------------------------------------------------
# Блок 4 — Таблица расходов проекта (только если есть)
# ---------------------------------------------------------------------------
project_expenses: list[dict] = list(r.get("project_expenses", []))
expenses_df = pd.DataFrame()
if project_expenses:
    st.markdown("#### Расходы проекта")
    exp_rows = []
    for e in project_expenses:
        category = e.get("category", "")
        amount = cx(float(e.get("amount", 0.0)))
        exp_rows.append(
            {
                "Название": e.get("name", ""),
                "Тип": EXPENSE_CATEGORY_LABELS.get(category, category),
                f"Сумма, {sym}": f"{amount:,.0f}",
                "Перевыставляется клиенту": "Да" if e.get("billable") else "Нет",
            }
        )
    expenses_df = pd.DataFrame(exp_rows)
    st.dataframe(expenses_df, width='stretch', hide_index=True)


st.markdown("---")


# ---------------------------------------------------------------------------
# Блок 5 — Индикатор рентабельности + экспорт
# ---------------------------------------------------------------------------
nne_raw = float(r["nne"])  # знак не зависит от конвертации валюты
if nne_raw > 0:
    st.success("✅ Проект рентабелен")
elif nne_raw == 0:
    st.warning("⚠️ Проект в точке безубыточности")
else:
    st.error("❌ Проект убыточен — пересмотрите ставки или состав команды")


# --- Скачать смету (XLSX) ---
# Формируем .xlsx через pandas + openpyxl. Два листа:
#  1) «Смета по этапам» — заголовок с названием проекта и датой + таблица этапов
#     + последняя строка «Итого» жирным.
#  2) «Финансовые показатели» — плоский список «показатель: значение».
# Весь файл собирается в памяти (BytesIO), чтобы отдать st.download_button.
def _build_xlsx_report() -> bytes:
    """Собирает xlsx-отчёт по текущей смете и возвращает его как bytes.

    Логика вынесена в функцию, чтобы не засорять основной поток рендера
    и чтобы её можно было при желании вызвать в тесте.
    """
    output = io.BytesIO()
    today_str = date.today().isoformat()
    project_title = (r.get("project_name") or "project").strip() or "project"

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # --- Лист 1: Смета по этапам ---
        sheet1 = "Смета по этапам"
        if stages_df.empty:
            # Пустой DataFrame всё равно нужно записать, чтобы лист существовал.
            pd.DataFrame([{"Примечание": "В смете нет этапов"}]).to_excel(
                writer, sheet_name=sheet1, index=False
            )
        else:
            # Шапка: строка 1 — заголовок (4 ячейки объединим), строка 2 — пустая,
            # со строки 3 — таблица. startrow=2 пишет начиная с excel-row 3.
            stages_df.to_excel(writer, sheet_name=sheet1, index=False, startrow=2)

            ws = writer.sheets[sheet1]
            # Объединяем ячейки A1:F1 (шесть столбцов таблицы) и ставим заголовок.
            n_cols = len(stages_df.columns)
            end_col_letter = ws.cell(row=1, column=n_cols).column_letter
            ws.merge_cells(f"A1:{end_col_letter}1")
            header_cell = ws["A1"]
            header_cell.value = f"{project_title} — смета от {today_str}"
            # Жирный заголовок (openpyxl.styles.Font импортируем лениво).
            from openpyxl.styles import Font

            header_cell.font = Font(bold=True, size=12)

            # Жирная последняя строка (итоги): учитываем startrow=2 и шапку
            # таблицы → Excel-строка = 3 (шапка) + len(stages_df) (данные) = total.
            total_row_excel = 3 + len(stages_df)
            for col_idx in range(1, n_cols + 1):
                ws.cell(row=total_row_excel, column=col_idx).font = Font(bold=True)

        # --- Лист 2: Финансовые показатели ---
        sheet2 = "Финансовые показатели"
        currency_label = f" ({sym})"
        # Показатели дублируют карточки дашборда, но в плоском виде.
        margin_pct = (nne_val / gross * 100) if gross else 0.0
        effective_tax_pct = (cx(r["tax"]) / cx(r["gross_revenue"]) * 100) if r["gross_revenue"] else 0.0

        fin_rows = [
            {"Показатель": "Выручка (без перевыставляемых)" + currency_label, "Значение": f"{gross:,.0f}"},
            {"Показатель": "Перевыставляемые расходы" + currency_label, "Значение": f"{disb_b:,.0f}"},
            {"Показатель": "Итоговый счёт клиенту" + currency_label, "Значение": f"{total_client:,.0f}"},
            {"Показатель": "Прямые трудозатраты" + currency_label, "Значение": f"{cx(r['direct_labor']):,.0f}"},
            {"Показатель": "Накладные (распределённые)" + currency_label, "Значение": f"{cx(r['overheads_alloc']):,.0f}"},
            {"Показатель": "Расходы за счёт фирмы" + currency_label, "Значение": f"{cx(r['disbursements_own']):,.0f}"},
            {"Показатель": "Налог" + currency_label, "Значение": f"{cx(r['tax']):,.0f}"},
            {"Показатель": "Чистая прибыль (NNE)" + currency_label, "Значение": f"{nne_val:,.0f}"},
            {"Показатель": "Маржа, %", "Значение": f"{margin_pct:.1f}"},
            {"Показатель": f"Средневзвешенная ставка ({sym}/ч)", "Значение": f"{cx(r['blended_rate']):,.0f}"},
            {"Показатель": "Коэффициент рычага", "Значение": f"{r['leverage']:.2f}"},
            {"Показатель": "Эффективная ставка налога, %", "Значение": f"{effective_tax_pct:.1f}"},
            {"Показатель": "Налоговый режим", "Значение": r.get("regime_label", "—")},
        ]
        pd.DataFrame(fin_rows).to_excel(writer, sheet_name=sheet2, index=False)

    return output.getvalue()


try:
    xlsx_bytes = _build_xlsx_report()
    file_name = f"Rentab_смета_{(r.get('project_name') or 'project').strip() or 'project'}_{date.today().isoformat()}.xlsx"
    st.download_button(
        label="📥 Скачать смету (Excel)",
        data=xlsx_bytes,
        file_name=file_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        width='content',
    )
except ImportError as exc:
    # openpyxl может отсутствовать в окружении — мягко деградируем.
    st.warning(
        f"Экспорт в Excel недоступен: не установлен openpyxl ({exc}). "
        "Добавьте его в requirements.txt и перезапустите приложение."
    )
